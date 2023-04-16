from openpyxl import Workbook
from openpyxl import load_workbook
import os
wb = load_workbook("sample.xlsx")

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime

ws['A3'] = '=EXTRAE(D5,1,ENCONTRAR(".",D5,1)-1)'

# Save the file
wb.close()



# se creo el archivo.csv y se colocaron dos columnas con nombre (Nombre,Hora) para despues leer el archivo
persona='=EXTRAE(D5,1,ENCONTRAR(".",D5,1)-1)'
f = open('sample.xlsx', 'r+')  # se lee el archivo
lista_datos = f.readlines()  # se len la cantidad de columndas
nombre_registro = []
for linea in lista_datos:  # se quenea un lista de datos apartir de las columnas
    ingreso = linea.split(',')  # se genera un espacio
    print(ingreso)
    nombre_registro.append(ingreso[0])  # se ingresa el nombre del usuario
    print(nombre_registro)
if persona not in nombre_registro:
    ahora = datetime.now()  # se ingresa la fecha de la persona
    string_ahora = ahora.strftime('%H:%M:%S')
    f.writelines(f'\n{persona}, {string_ahora}')
f.close()